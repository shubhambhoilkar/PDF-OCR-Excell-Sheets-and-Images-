# Option 1 — Two zones (LEFT + RIGHT) + full-width tables.

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

PDF_PATH = r"Payslip_Prakash.pdf"
OUTPUT_XLSX = "structured-output\\new1_Payslip_Prakash.xlsx"

# Config:
RIGHT_ZONE_START_COL = 16   # Excel column index where the right zone begins (1-based)
H_SPACING_ROWS = 1          # blank rows between stacked tables within same zone
FULL_WIDTH_THRESHOLD = 0.6  # if table width / page_width >= this -> treat as full-width
BORDER_SIDE = Side(style="thin")
BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

def write_table(ws, table_rows, start_row, start_col):
    """Write table (list of lists) at given start_row, start_col. Returns end_row."""
    r = start_row
    max_col = start_col
    for row in table_rows:
        c = start_col
        for cell in row:
            val = cell if (cell is not None) else ""
            ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            c += 1
        if c - 1 > max_col:
            max_col = c - 1
        r += 1
    end_row = r - 1
    # Apply border only around the table region
    for rr in range(start_row, end_row + 1):
        for cc in range(start_col, max_col + 1):
            ws.cell(row=rr, column=cc).border = BORDER
    return end_row

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "TwoZones"

    left_row_ptr = 1
    right_row_ptr = 1
    full_row_ptr = 1  # we will place full-width tables using max of both pointers

    with pdfplumber.open(PDF_PATH) as pdf:
        page = pdf.pages[0]  # adjust to iterate pages if needed
        page_width = page.width

        # find_tables returns Table objects with bbox and .extract()
        tables = list(page.find_tables())

        for t in tables:
            try:
                bbox = t.bbox  # (x0, top, x1, bottom)
            except Exception:
                # fallback: we don't have bbox, treat as full-width
                bbox = (0, 0, page_width, 0)

            x0, _, x1, _ = bbox
            table_width = (x1 - x0) if (x1 > x0) else page_width

            # extract rows (list of lists)
            rows = t.extract()

            # decide placement
            if table_width / page_width >= FULL_WIDTH_THRESHOLD:
                # full width: place below the current max of both columns
                place_row = max(left_row_ptr, right_row_ptr, full_row_ptr)
                end_row = write_table(ws, rows, place_row, 1)
                # update both pointers so subsequent left/right don't overlap
                left_row_ptr = end_row + H_SPACING_ROWS + 1
                right_row_ptr = end_row + H_SPACING_ROWS + 1
                full_row_ptr = end_row + H_SPACING_ROWS + 1
            else:
                # choose left or right by x0 (left half -> left zone)
                if x0 < (page_width / 2):
                    end_row = write_table(ws, rows, left_row_ptr, 1)
                    left_row_ptr = end_row + H_SPACING_ROWS + 1
                else:
                    end_row = write_table(ws, rows, right_row_ptr, RIGHT_ZONE_START_COL)
                    right_row_ptr = end_row + H_SPACING_ROWS + 1

    # Optional: autofit a reasonable width for columns used
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    wb.save(OUTPUT_XLSX)
    print("Saved:", OUTPUT_XLSX)

if __name__ == "__main__":
    main()

