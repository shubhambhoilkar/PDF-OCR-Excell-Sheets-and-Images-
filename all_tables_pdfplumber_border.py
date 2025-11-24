import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ================================
# CONFIG
# ================================
pdf_path = "Payslip.pdf"
output_excel = "Payslip_all_tables_borders.xlsx"


# ================================
# STEP 1 — Extract all tables
# ================================
all_tables = []

with pdfplumber.open(pdf_path) as pdf:
    for page in pdf.pages:
        page_tables = page.extract_tables()
        if page_tables:
            all_tables.extend(page_tables)


# ================================
# STEP 2 — Create Excel workbook
# ================================
wb = Workbook()
ws = wb.active
ws.title = "PayslipRaw"

thin_border = Border(
    left= Side(style= 'thin'),
    right= Side(style= 'thin'),
    top= Side(style= 'thin'),
    bottom= Side(style= 'thin')
)
current_row = 1

for table in all_tables:
    # Convert table to DataFrame
    df = pd.DataFrame(table)

    start_row = current_row

    # Write DataFrame rows to the sheet
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
        current_row += 1

    # Apply Borders
    end_rows = current_row - 1
    max_col = ws.max_column

    for r in range(start_row, end_rows + 1):
        for c in range(1, max_col +1 ):
            cell = ws.cell(row = r, column= c)
            cell.border =  thin_border


    # Add two blank rows for separation
    # current_row += len(df) + 2
    ws.append([])
    ws.append([])
    current_row += 2


# ================================
# STEP 3 — Save Excel
# ================================
wb.save(output_excel)

print(f"Excel created successfully: {output_excel}")

