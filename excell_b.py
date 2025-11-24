# pdf_to_excel_tables_placed.py
# Recreate only table portions exactly (tables + layout surrounding kept simple)
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

PDF_PATH = r"C:\\Users\\Developer\\Shubham_files\\PDF-OCR\\Payslip.pdf"
OUTPUT_XLSX = "Payslip_Tables_Placed_B.xlsx"

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "TablesOnly"

    with pdfplumber.open(PDF_PATH) as pdf:
        current_row = 1
        for page in pdf.pages:
            # pdfplumber's extract_tables returns lists-of-lists (rows)
            tables = page.extract_tables()
            # also get approximate top positions per table via find_table? pdfplumber doesn't provide
            # positions from extract_tables, so we will just stack them with a small spacing to match order.
            for table in tables:
                if not table:
                    continue
                df = pd.DataFrame(table)
                # write df rows maintaining cells
                for row in dataframe_to_rows(df, index=False, header=False):
                    ws.append(row)
                    current_row += 1
                # apply borders for recently added rows
                start = current_row - len(df)
                end = current_row - 1
                max_col = ws.max_column
                for r in range(start, end + 1):
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                # add spacing
                ws.append([])
                ws.append([])
                current_row += 2

    wb.save(OUTPUT_XLSX)
    print("Saved:", OUTPUT_XLSX)

if __name__ == "__main__":
    main()
