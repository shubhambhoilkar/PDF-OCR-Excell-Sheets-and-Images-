# pdf_to_excel_tables_structured.py
# Exact table structures (no decorative layout)

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

PDF_PATH = r"Payslip.pdf"
OUTPUT_XLSX = "Payslip_Tables_Structured_C.xlsx"

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def autofit_col_widths(ws, start_row=1):
    from openpyxl.utils import get_column_letter
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(start_row, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            # consider string length; adjust factor if needed
            l = len(str(val))
            if l > max_len:
                max_len = l
        # set width with a padding
        ws.column_dimensions[get_column_letter(col)].width = max(10, max_len * 1.1)

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "StructuredTables"

    with pdfplumber.open(PDF_PATH) as pdf:
        current_row = 1
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                df = pd.DataFrame(table)
                # write table rows
                start_row = current_row
                for row in dataframe_to_rows(df, index=False, header=False):
                    ws.append(row)
                    current_row += 1
                end_row = current_row - 1
                # style and border
                max_col = ws.max_column
                for r in range(start_row, end_row + 1):
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                # small spacing
                ws.append([])
                current_row += 1

    # autofit columns based on content
    autofit_col_widths(ws)
    wb.save(OUTPUT_XLSX)
    print("Saved:", OUTPUT_XLSX)

if __name__ == "__main__":
    main()

