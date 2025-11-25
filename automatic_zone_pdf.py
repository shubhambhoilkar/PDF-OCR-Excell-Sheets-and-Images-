"""
Option 3 — Automatic placement:
- full-width tables (span large portion)
- narrow tables placed left/center/right based on x0
"""

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
import math

PDF_PATH = r"C:\\Users\\Developer\\Shubham_files\\PDF-OCR\\Payslip_Prakash.pdf"
OUTPUT_XLSX = "C:\\Users\\Developer\\Shubham_files\\Pdf_OCR\\structured-output\\new3_Payslip_Prakash.xlsx"

# Config
ZONE_COUNT = 3                # try 3 vertical zones (left, center, right)
ZONE_START_COLS = [1, 8, 15]  # default start columns for each zone
FULL_WIDTH_THRESHOLD = 0.55   # if width/page_width >= this => full-width
H_SPACING_ROWS = 1
BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

def write_table(ws, rows, start_row, start_col):
    r = start_row
    max_col = start_col
    for row in rows:
        c = start_col
        for cell in row:
            val = cell if cell is not None else ""
            ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            c += 1
        if c - 1 > max_col:
            max_col = c - 1
        r += 1
    end_row = r - 1
    for rr in range(start_row, end_row + 1):
        for cc in range(start_col, max_col + 1):
            ws.cell(row=rr, column=cc).border = BORDER
    return end_row

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "AutoZones"

    zone_rows = [1] * ZONE_COUNT
    full_row_ptr = 1

    with pdfplumber.open(PDF_PATH) as pdf:
        page = pdf.pages[0]
        page_width = page.width
        tables = list(page.find_tables())

        for t in tables:
            bbox = getattr(t, "bbox", (0,0,page_width,0))
            x0, _, x1, _ = bbox
            table_width = (x1 - x0) if (x1 > x0) else page_width
            rows = t.extract()

            # classify
            if table_width / page_width >= FULL_WIDTH_THRESHOLD:
                place_row = max(full_row_ptr, *zone_rows)
                end = write_table(ws, rows, place_row, 1)
                # update pointers
                full_row_ptr = end + H_SPACING_ROWS + 1
                for i in range(len(zone_rows)):
                    zone_rows[i] = full_row_ptr
            else:
                # decide which zone by x0 proportion
                proportion = x0 / max(1.0, page_width)
                # map proportion to zone index
                zone_idx = min(ZONE_COUNT - 1, int(proportion * ZONE_COUNT))
                start_col = ZONE_START_COLS[zone_idx] if zone_idx < len(ZONE_START_COLS) else (1 + zone_idx * 7)
                end = write_table(ws, rows, zone_rows[zone_idx], start_col)
                zone_rows[zone_idx] = end + H_SPACING_ROWS + 1

    # autofit reasonable column widths
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    wb.save(OUTPUT_XLSX)
    print("Saved:", OUTPUT_XLSX)

if __name__ == "__main__":
    main()
