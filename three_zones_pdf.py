# Option 2 — Three zones (LEFT + CENTER + RIGHT) + full-width tables.

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

PDF_PATH = r"C:\\Users\\Developer\\Shubham_files\\PDF-OCR\\Payslip_Prakash.pdf"
OUTPUT_XLSX = "C:\\Users\\Developer\\Shubham_files\\Pdf_OCR\\structured-output\\new2_Payslip_Prakash.xlsx"

# Config:
LEFT_START_COL = 1
CENTER_START_COL = 8
RIGHT_START_COL = 15
H_SPACING_ROWS = 1
FULL_WIDTH_THRESHOLD = 0.6
BORDER_SIDE = Side(style="thin")
BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

def write_table(ws, table_rows, start_row, start_col):
    r = start_row
    max_col = start_col
    for row in table_rows:
        c = start_col
        for cell in row:
            val = cell if (cell is not None) else ""
            ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            c += 1
        if c - 1 > max_col:
            max_col = c - 1
        r += 1
    end_row = r - 1
    # apply border only on table
    for rr in range(start_row, end_row + 1):
        for cc in range(start_col, max_col + 1):
            ws.cell(row=rr, column=cc).border = BORDER
    return end_row

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "ThreeZones"

    left_row = 1
    center_row = 1
    right_row = 1
    full_row_ptr = 1

    with pdfplumber.open(PDF_PATH) as pdf:
        page = pdf.pages[0]
        page_width = page.width
        tables = list(page.find_tables())

        for t in tables:
            bbox = getattr(t, "bbox", (0,0,page_width,0))
            x0, _, x1, _ = bbox
            table_width = (x1 - x0) if (x1 > x0) else page_width
            rows = t.extract()

            if table_width / page_width >= FULL_WIDTH_THRESHOLD:
                place_row = max(left_row, center_row, right_row, full_row_ptr)
                end = write_table(ws, rows, place_row, LEFT_START_COL)
                left_row = end + H_SPACING_ROWS + 1
                center_row = end + H_SPACING_ROWS + 1
                right_row = end + H_SPACING_ROWS + 1
                full_row_ptr = end + H_SPACING_ROWS + 1
            else:
                # decide zone: left third, center third, right third
                third = page_width / 3
                if x0 < third:
                    end = write_table(ws, rows, left_row, LEFT_START_COL)
                    left_row = end + H_SPACING_ROWS + 1
                elif x0 < 2 * third:
                    end = write_table(ws, rows, center_row, CENTER_START_COL)
                    center_row = end + H_SPACING_ROWS + 1
                else:
                    end = write_table(ws, rows, right_row, RIGHT_START_COL)
                    right_row = end + H_SPACING_ROWS + 1

    # tidy column widths
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    wb.save(OUTPUT_XLSX)
    print("Saved:", OUTPUT_XLSX)

if __name__ == "__main__":
    main()
