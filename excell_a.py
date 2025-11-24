# pdf_to_excel_layout_full.py
import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
import math
import statistics

PDF_PATH = r"Payslip.pdf"
OUTPUT_XLSX = "Payslip_Layout_Full_Reconstruction_A.xlsx"

# parameters you can tune
Y_CLUSTER_TOL = 3      # group words whose 'top' is within this many pts
COL_SPLIT_GAP = 10     # minimum gap (pts) to consider a new "visual column"
EXCEL_PIXELS_PER_CHAR = 7  # approx mapping; adjust if column widths look weird

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cluster_rows(words):
    # words: list of dicts with 'text','x0','x1','top','bottom'
    ys = sorted(words, key=lambda w: w['top'])
    clusters = []
    for w in ys:
        placed = False
        for c in clusters:
            # use top difference for clustering
            if abs(c['y_mean'] - w['top']) <= Y_CLUSTER_TOL:
                c['words'].append(w)
                c['y_vals'].append(w['top'])
                c['y_mean'] = statistics.mean(c['y_vals'])
                placed = True
                break
        if not placed:
            clusters.append({'words':[w], 'y_vals':[w['top']], 'y_mean':w['top']})
    # sort clusters top->bottom
    clusters.sort(key=lambda c: c['y_mean'])
    return clusters

def compute_columns(words):
    # determine vertical column boundaries from all x0/x1 spans
    edges = []
    for w in words:
        edges.append(w['x0'])
        edges.append(w['x1'])
    edges = sorted(set(edges))
    # compress edges: if gaps between adjacent edges are small, merge
    cols = [edges[0]]
    for e in edges[1:]:
        if e - cols[-1] > COL_SPLIT_GAP:
            cols.append(e)
    # append last boundary slightly beyond max x1
    cols.append(max(edges) + COL_SPLIT_GAP)
    return cols  # list of x positions representing left boundaries

def find_col_index(x, columns):
    # columns are left boundaries; last element is rightmost boundary
    for i in range(len(columns)-1):
        if x >= columns[i] and x < columns[i+1]:
            return i
    return len(columns)-2

def span_columns(x0, x1, columns):
    left = find_col_index(x0, columns)
    right = find_col_index(x1 - 1e-6, columns)
    return left, right

def pts_to_excel_colwidth(pts):
    # very rough conversion: 1 Excel width unit ≈ 7 pixels; PDF pts ~ pixels at 72dpi
    # adjust using EXCEL_PIXELS_PER_CHAR if needed
    pixels = pts  # assume 1pt ~ 1px here; tweak if different
    return max(3, pixels / EXCEL_PIXELS_PER_CHAR)

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "LayoutFull"

    all_words = []
    with pdfplumber.open(PDF_PATH) as pdf:
        page = pdf.pages[0]  # single page payslip; if multi-page, loop pages
        # extract positioned words
        words = page.extract_words(use_text_flow=True, keep_blank_chars=False, x_tolerance=1, y_tolerance=1)
        # standardize keys
        words2 = []
        for w in words:
            words2.append({
                "text": w.get("text",""),
                "x0": float(w.get("x0",0)),
                "x1": float(w.get("x1",0)),
                "top": float(w.get("top",0)),
                "bottom": float(w.get("bottom",0))
            })
        all_words = words2

    # cluster rows by y coordinate
    clusters = cluster_rows(all_words)

    # compute column vertical boundaries across whole page
    columns = compute_columns(all_words)

    # prepare a grid: rows × columns
    grid = []
    merges = []  # list of merges as (r1, c1, r2, c2)
    for row_idx, c in enumerate(clusters, start=1):
        # initialize empty cells for this row
        row_cells = [""] * (len(columns)-1)
        grid.append(row_cells)
        # sort words left->right
        items = sorted(c['words'], key=lambda w: w['x0'])
        for w in items:
            left, right = span_columns(w['x0'], w['x1'], columns)
            # if span > 0, mark merge
            if left <= right:
                # accumulate text into the left-most cell
                existing = row_cells[left]
                if existing:
                    row_cells[left] = existing + " " + w['text']
                else:
                    row_cells[left] = w['text']
                if right > left:
                    merges.append((row_idx, left+1, row_idx, right+1))  # excel cols 1-indexed
            else:
                # fallback assign to left
                idx = max(0, left)
                row_cells[idx] = (row_cells[idx] + " " + w['text']).strip()

    # write to worksheet
    for r_index, row in enumerate(grid, start=1):
        for c_index, val in enumerate(row, start=1):
            if val is None:
                val = ""
            ws.cell(row=r_index, column=c_index, value=val)
            ws.cell(row=r_index, column=c_index).alignment = Alignment(wrap_text=True, vertical="top")

    # apply merges
    for (r1, c1, r2, c2) in merges:
        try:
            # only merge if range has multiple columns
            if c2 > c1:
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        except Exception:
            pass

    # set column widths based on column boundaries spacing
    col_widths = []
    for i in range(len(columns)-1):
        width_pts = columns[i+1] - columns[i]
        col_widths.append(pts_to_excel_colwidth(width_pts))

    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # apply borders across used region
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            cell.border = border

    wb.save(OUTPUT_XLSX)
    print("Saved:", OUTPUT_XLSX)

if __name__ == "__main__":
    main()

